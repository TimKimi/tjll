"""统一文档加载：按后缀返回纯文本 / Markdown。"""

from __future__ import annotations

from pathlib import Path

from backend.rag.document.pdf import parse_pdf_with_mineru


def load_document_as_text(file_path: str) -> str:
    """根据后缀选择 Loader，返回纯文本（PDF 经 MinerU 转为 Markdown）。"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    ext = path.suffix.lower()

    if ext == ".pdf":
        md_path = parse_pdf_with_mineru(str(path))
        return Path(md_path).read_text(encoding="utf-8")

    if ext == ".md":
        return path.read_text(encoding="utf-8")

    if ext == ".txt":
        from langchain_community.document_loaders import TextLoader

        return TextLoader(str(path), encoding="utf-8").load()[0].page_content

    if ext in (".doc", ".docx"):
        from langchain_community.document_loaders import (
            UnstructuredWordDocumentLoader,
        )

        docs = UnstructuredWordDocumentLoader(str(path)).load()
        return "\n\n".join(doc.page_content for doc in docs)

    if ext in (".xls", ".xlsx"):
        return _load_excel_as_text(path)

    raise ValueError(
        f"不支持的文件类型: {ext}。支持: .pdf .md .txt .doc .docx .xls .xlsx"
    )


def _load_excel_as_text(path: Path) -> str:
    """将 Excel 各 sheet 拼成文本。优先 Unstructured，回退 openpyxl。"""
    try:
        from langchain_community.document_loaders import UnstructuredExcelLoader

        docs = UnstructuredExcelLoader(str(path)).load()
        return "\n\n".join(doc.page_content for doc in docs)
    except Exception:
        pass

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "读取 Excel 需要 unstructured 或 openpyxl，请先安装依赖。"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)
